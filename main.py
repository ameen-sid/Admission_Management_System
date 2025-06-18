# Importing Libraries
from tkinter import *
from tkinter import messagebox
from openpyxl import *
# import xlrd
import pathlib
import time


# Splash Screen Class
class GUI(Tk):
    def __init__(self):
        super().__init__()

    def splash(self):
        self.width = 340
        self.height = 200
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = (self.screen_width / 2) - (self.width / 2)
        self.y = (self.screen_height / 2) - (self.height / 2)
        self.geometry("%dx%d+%d+%d" %(self.width, self.height, self.x, self.y))
        self.overrideredirect(1)
    
    def frame(self):
        self.frame_1 = Frame(width=340, height=200, bg="#ffffff").place(x=0, y=0)
        self.main_heading = Label(self.frame_1, text="AMEEN SID", fg="#000000", bg="#ffffff", font="Elianto 24 normal").place(x=75, y=70)
        self.sub_heading = Label(self.frame_1, text="Present", fg="#000000", bg="#ffffff", font="Calibri 10 normal", padx=0, pady=0).place(x=220, y=105)
    
    def loading(self):
        for i in range(1):
            l1 = Label(text="_", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="__________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="____________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="______________________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="________________________________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="__________________________________________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="____________________________________________________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)

            l1 = Label(text="______________________________________________________________________", fg="#000000", bg="#ffffff").place(x=0, y=183)
            self.update()
            time.sleep(0.5)


# Main Screen Class
class Root(Tk):
    def __init__(self):
        super().__init__()
        self.state('zoomed')
        self.resizable(False, False)
        self.title("Ameen Sid")
        # self.iconphoto(0, "/Project/school.png")
        # self.wm_iconbitmap("school.ico")
        # self.minsize(1366, 768)
        # self.geometry("720x520")
        # self.s = Style(self)
        # self.s.theme_use(clam)
        self.config(bg="#ffffff") #171F24 #121212
        
    def menus(self):
        # Functions
            # File Menu's Functions
        def new_form():
            self.class_entry.delete(0, 'end')
            self.first_entry.delete(0, 'end')
            self.last_entry.delete(0, 'end')
            self.year_entry.delete(0, 'end')
            self.parent_entry.delete(0, 'end')
            self.address_entry.delete('0', 'end')
            self.address_line2_entry.delete(0, 'end')
            self.city_entry.delete(0, 'end')
            self.pincode_entry.delete(0, 'end')
            self.email_entry.delete(0, 'end')
            self.phone_entry.delete(0, 'end')

        def open_record():
            def getting_id():
                show_window = Tk()
                show_window.title("Student Record")
                show_window.resizable(0, 0)
                show_window.geometry("400x400")
                show_window.config(bg="#ffffff")
                # Getting ID from input field
                # print(id_entry.get()[-1])
                l = Label(show_window, bg="#ffffff", fg="#000000", font="Calibri 18 normal")
                l.pack(anchor='center', pady=20)
                file = load_workbook('Data.xlsx')
                sheet = file.active
                # print(sheet.max_row)
                # print(id_entry.get())
                if len(id_entry.get()) < 9:
                    # data = sheet['A'+str(id_entry.get()[-1])]
                    data = sheet[id_entry.get()[-1]]
                    if data[0].value == "Cancelled":
                        l.config(text="Invalid Student ID!!")
                    else:
                        if sheet.max_row >= int(id_entry.get()[-1]):
                            #TODO Fix this bug (max row is increasing)
                            # print(sheet.max_row)
                            # print(int(id_entry.get()))
                            # print(int(id_entry.get()[-1]))
                            data = sheet[id_entry.get()[-1]]
                            list = ""
                            for i in data:
                                list = f"{list + str(i.value)}\n"
                            l.config(text=list)
                        else:
                            print(here)
                            l.config(text="Invalid Student ID!")
                else:
                    data = sheet['A'+str(id_entry.get()[-2:])]
                    if data.value == "Cancelled":
                        l.config(text="Invalid Student ID!!")
                    else:
                        if sheet.max_row >= int(id_entry.get()[-2:]):
                            data = sheet[id_entry.get()[-2:]]
                            list = ""
                            for i in data:
                                list = f"{list + str(i.value)}\n"
                            l.config(text=list)
                        else:
                            l.config(text="Invalid Student ID!")
                show_window.mainloop()
            
            record_window = Tk()
            record_window.title("Student Record")
            record_window.resizable(0, 0)
            record_window.geometry("480x200")
            record_window.config(bg="#ffffff")
            # Getting Student ID
            Label(record_window, text="Enter your Student ID :", bg="#ffffff", fg="#000000", font="Calibri 20 bold").pack(anchor='center', pady=20)
                # Variable
            _id = StringVar()
            id_entry = Entry(record_window, textvariable=_id, font="Calibri 20 bold", bg="#ffffff", fg="#000000")
            id_entry.pack(anchor='center')
            Button(record_window, text="Submit", command=getting_id, bg="#ffffff", fg="#000000", font="Calibri 20 bold").pack(anchor='center', pady=5)
            record_window.mainloop()

        def edit_admission_details():
            # Edit Button's Function
            def edit():
                if (student_id.get() == "") or (edited_input.get() == ""):
                    messagebox.showerror("Error", "Please fill the field!")
                else:
                    # Open File
                    file = load_workbook('Data.xlsx')
                    sheet = file.active
                    # Full Name
                    if option_click.get() == "Full Name":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=2).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=2).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Date of Birth
                    elif option_click.get() == "Date of Birth":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=3).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=3).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Parent/Guardian Name
                    elif option_click.get() == "Parent/Guardian Name":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=4).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=4).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Full Address
                    elif option_click.get() == "Full Address":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=5).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=5).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # City
                    elif option_click.get() == "City":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=6).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=6).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Pincode
                    elif option_click.get() == "Pincode":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=7).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=7).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Email
                    elif option_click.get() == "Email":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=8).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=8).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    # Phone
                    elif option_click.get() == "Phone":
                        if len(student_id.get()) < 9:
                            sheet.cell(row=int(student_id.get()[-1]), column=9).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                        else:
                            sheet.cell(row=int(student_id.get()[-2:]), column=9).value=edited_input.get()
                            messagebox.showinfo("Information", "Your Details Updated Successfully!")
                    edit_details_window.destroy()
                    # Save File
                    file.save("Data.xlsx")

            edit_details_window = Tk()
            edit_details_window.title("Edit Your Personal Details")
            edit_details_window.resizable(0, 0)
            edit_details_window.geometry("600x600")
            edit_details_window.config(bg="#ffffff")
            # Variable
            option_click = StringVar()
            new_detail = StringVar()
            student_id_var = Variable()
            # Heading
            Label(edit_details_window, text="Edit Your Personal Details", font="Calibri 25 underline bold", bg="#ffffff", fg="#000000").pack(anchor='center', pady=15)
            # Label for Student ID
            Label(edit_details_window, text="Enter Your Student ID :-", font="Calibri 20 bold", bg="#ffffff", fg="#000000").pack(anchor='center', pady=15)
            # Getting Student ID
            student_id = Entry(edit_details_window, textvariable=student_id_var, font="Calibri 20 bold", fg="#000000", bg="#ffffff")
            student_id.pack(anchor='center', pady=15)
            # Label for Select Option
            Label(edit_details_window, text="Choose Option to Edit Detail :-", font="Calibri 20 bold", bg="#ffffff", fg="#000000").pack(anchor='center')
            option_list = ["Full Name", "Date of Birth", "Parent/Guardian Name", "Full Address", "City", "Pincode", "Email", "Phone"]
            option_click.set(option_list[0])            
            section_drop = OptionMenu(edit_details_window, option_click, *option_list).pack(anchor='center', pady=10)
            # Input Entry
            edited_input = Entry(edit_details_window, textvariable=new_detail, font="Calibri 20 bold", fg="#000000", bg="#ffffff")
            edited_input.pack(anchor='center', pady=15)
            # Edit Button
            Button(edit_details_window, text="Edit Detail", font="Calibri 20 bold", fg="#000000", bg="#ffffff", command=edit).pack(anchor='center', pady=10)
            edit_details_window.mainloop()

        def cancel_admission():
            # Cancel Button's Function
            def cancel():
                if student_id_entry.get() == "":
                    messagebox.showerror("Error", "please fill the field")
                else:
                    ans = messagebox.askquestion("Confirm", "Are sure you want to cancel admission?")
                    if ans == "yes":
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        for i in range(1, 10):
                            if len(student_id_entry.get()) < 9:
                                sheet.cell(row=int(student_id_entry.get()[-1]), column=i).value="Cancelled"
                            else:
                                sheet.cell(row=int(student_id_entry.get()[-2:]), column=i).value="Cancelled"
                        messagebox.showinfo("Information", "Your Admission is Canceled!")
                        # if len(student_id_entry.get()) < 9:
                        #     # sheet.delete_rows(int(student_id_entry.get()[-1]), 1)
                        # else:
                        #     # sheet.delete_rows(int(student_id_entry.get()[-2:]), 1)
                        file.save("Data.xlsx")                        
                    else:
                        pass

            cancel_window = Tk()
            cancel_window.title("Cancel Admission")
            cancel_window.resizable(0, 0)
            cancel_window.geometry("600x300")
            cancel_window.config(bg="#ffffff")
            # ID Variable
            student_id = StringVar()
            # Heading
            heading = Label(cancel_window, text="Cancel Admission", fg="#000000", bg="#ffffff", font="Calibri 25 underline bold")
            heading.pack(anchor='center', pady=20)
            # Student ID Label
            student_id_label = Label(cancel_window, text="Student ID :-", fg="#000000", bg="#ffffff", font="Calibri 15 normal")
            student_id_label.place(x=100, y=100)
            # Student ID Entry Field
            student_id_entry = Entry(cancel_window, textvariable=student_id, font="Calibri 15 normal", fg="#000000", bg="#ffffff")
            student_id_entry.place(x=220, y=103)
            # Cancel Button
            cancel_button = Button(cancel_window, text="Cancel Admission", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=cancel)
            cancel_button.place(x=200, y=150)
            cancel_window.mainloop()
        
            # Student Zone's Menu Functions
        def fee_status():
            # Check Status Button's Function
            def check_status():
                if student_id_entry.get() == "":
                    messagebox.showerror("Error", "Please fill the field!")
                else:
                    file = load_workbook('Data.xlsx')
                    sheet = file.active
                    if len(student_id_entry.get()) < 9:
                        print(sheet['J'+str(student_id_entry.get()[-1])].value)
                        if sheet['J'+str(student_id_entry.get()[-1])].value == "Submitted":
                            messagebox.showinfo("Information", "Your fees is submitted!")
                        else:
                            messagebox.showerror("Error", "Your fees is not submitted!")
                    else:
                        if sheet['J'+str(student_id_entry.get()[-2:])].value == "Submitted":
                            messagebox.showinfo("Information", "Your fees is submitted!")
                        else:
                            messagebox.showerror("Error", "Your fees is not submitted!")
                        # messagebox.showinfo("Information", "Your fees Submitted Successfully!")
                        status_window.destroy()
                        file.save("Data.xlsx")

            status_window = Tk()
            status_window.title("Fee Status")
            status_window.resizable(0, 0)
            status_window.geometry("300x300")
            status_window.config(bg="#ffffff")
            # Variable
            student_id_var = Variable()
            # Heading
            Label(status_window, text="Fee Status", font="Calibri 30 underline bold", fg="#000000", bg="#ffffff").pack(anchor='center', pady=20)
            # Label for Student ID
            Label(status_window, text="Student ID :-", bg="#ffffff", fg="#000000", font="Calibri 20 bold").pack(anchor='center', pady=5)
            # Student ID Entry Field
            student_id_entry = Entry(status_window, textvariable=student_id_var, fg="#000000", bg="#ffffff", font="Calibri 20 bold")
            student_id_entry.pack(anchor='center', pady=5)
            # Check Status Button
            Button(status_window, text="Check Status", fg="#000000", bg="#ffffff", font="Calibri 20 bold", command=check_status).pack(anchor='center', pady=5)
            status_window.mainloop()

        def fee_payment():
            # Pay Button's Function
            def pay():
                if (payment_entry.get() == "") or (student_id_entry.get() == ""):
                    messagebox.showerror("Error", "Please fill the field!")
                else:
                    if len(payment_entry.get()) < 4:
                        messagebox.showerror("Error", "Enter Valid Amount!")
                    else:
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        if len(student_id_entry.get()) < 9:
                            sheet.cell(column=10, row=int(student_id_entry.get()[-1])).value="Submitted"
                        else:
                            sheet.cell(column=10, row=int(student_id_entry.get()[-2:])).value="Submitted"
                        messagebox.showinfo("Information", "Your fees Submitted Successfully!")
                        payment_window.destroy()
                        file.save("Data.xlsx")


            payment_window = Tk()
            payment_window.title("Fee Payment")
            payment_window.resizable(0, 0)
            payment_window.geometry("300x400")
            payment_window.config(bg="#ffffff")
            # Variable
            payment_var = Variable()
            student_id_var = Variable()
            # Heading
            Label(payment_window, text="Pay Your Fees", bg="#ffffff", fg="#000000", font="Calibri 30 underline bold").pack(anchor='center', pady=15)
            # Label for Student ID
            Label(payment_window, text="Student ID :-", bg="#ffffff", fg="#000000", font="Calibri 20 bold").pack(anchor='center', pady=5)
            # Student ID Entry Field
            student_id_entry = Entry(payment_window, textvariable=student_id_var, fg="#000000", bg="#ffffff", font="Calibri 20 bold")
            student_id_entry.pack(anchor='center', pady=5)
            # Label for Payment Amount
            Label(payment_window, text="Enter your fees amount :-", fg="#000000", bg="#ffffff", font="Calibri 20 bold").pack(anchor='center', pady=10)
            # Payment Entry Field
            payment_entry = Entry(payment_window, textvariable=payment_var, fg="#000000", bg="#ffffff", font="Calibri 20 bold")
            payment_entry.pack(anchor='center', pady=5)
            # Pay Button
            Button(payment_window, text="Pay", font="Calibri 20 bold", fg="#000000", bg="#ffffff", width=10, command=pay).pack(anchor='center', pady=10)
            payment_window.mainloop()

        def syllabus():
            pass

            # Edit Menu's Functions
        def capitalize_case():
            up = StringVar()
            up = self.focus_get().get()
            self.focus_get().delete(0, 'end')
            self.focus_get().insert(0, up.capitalize())

        def upper_case():
            up = StringVar()
            up = self.focus_get().get()
            self.focus_get().delete(0, 'end')
            self.focus_get().insert(0, up.upper())
        
        def lower_case():
            up = StringVar()
            up = self.focus_get().get()
            self.focus_get().delete(0, 'end')
            self.focus_get().insert(0, up.lower())

            # Find Menu's Functions
                # Find
        def find_in_fields():
            # Find Button's Function
            def find_in():
                if find_entry.get() == "":
                    messagebox.showerror("Error", "Please fill the field")
                else:
                    length_of_find_word = len(find_entry.get())
                    # checking in all fields
                    index_in_class = self.class_entry.get().find(find_entry.get())
                    index_in_first_name = self.first_entry.get().find(find_entry.get())
                    index_in_last_name = self.last_entry.get().find(find_entry.get())
                    index_in_year = self.year_entry.get().find(find_entry.get())
                    index_in_parent_name = self.parent_entry.get().find(find_entry.get())
                    index_in_address = self.address_entry.get().find(find_entry.get())
                    index_in_address_line2 = self.address_line2_entry.get().find(find_entry.get())
                    index_in_city = self.city_entry.get().find(find_entry.get())
                    index_in_pincode = self.pincode_entry.get().find(find_entry.get())
                    index_in_email = self.email_entry.get().find(find_entry.get())
                    index_in_phone = self.phone_entry.get().find(find_entry.get())
                    # Conditions
                    if (index_in_class < 0) and (index_in_first_name < 0) and (index_in_last_name < 0) and (index_in_year < 0) and (index_in_parent_name < 0) and (index_in_address < 0) and (index_in_address_line2 < 0) and (index_in_city < 0) and (index_in_pincode < 0) and (index_in_email < 0) and (index_in_phone < 0):
                        messagebox.showerror("Error", "Match Not Found!")
                    else:
                        messagebox.showinfo("Matched", "Match Found!")
                        find_window.destroy()
                        # Class Field
                        self.class_entry.select_range(index_in_class, index_in_class+length_of_find_word)
                        self.class_entry.icursor(index_in_class+length_of_find_word)
                        # First Name Field
                        self.first_entry.select_range(index_in_first_name, index_in_first_name+length_of_find_word)
                        self.first_entry.icursor(index_in_class+length_of_find_word)
                        # Last Name Field
                        self.last_entry.select_range(index_in_last_name, index_in_last_name+length_of_find_word)
                        self.last_entry.icursor(index_in_class+length_of_find_word)
                        # Year Field
                        self.year_entry.select_range(index_in_year, index_in_year+length_of_find_word)
                        self.year_entry.icursor(index_in_class+length_of_find_word)
                        # Parent Name Field
                        self.parent_entry.select_range(index_in_parent_name, index_in_parent_name+length_of_find_word)
                        self.parent_entry.icursor(index_in_class+length_of_find_word)
                        # Address Field
                        self.address_entry.select_range(index_in_address, index_in_address+length_of_find_word)
                        self.address_entry.icursor(index_in_class+length_of_find_word)
                        # Address Line 2 Field
                        self.address_line2_entry.select_range(index_in_address_line2, index_in_address_line2+length_of_find_word)
                        self.address_line2_entry.icursor(index_in_class+length_of_find_word)
                        # City Field
                        self.city_entry.select_range(index_in_city, index_in_city+length_of_find_word)
                        self.city_entry.icursor(index_in_class+length_of_find_word)
                        # Pincode Field
                        self.pincode_entry.select_range(index_in_pincode, index_in_pincode+length_of_find_word)
                        self.pincode_entry.icursor(index_in_class+length_of_find_word)
                        # Email Field
                        self.email_entry.select_range(index_in_email, index_in_email+length_of_find_word)
                        self.email_entry.icursor(index_in_class+length_of_find_word)
                        # Phone Field
                        self.phone_entry.select_range(index_in_phone, index_in_phone+length_of_find_word)
                        self.phone_entry.icursor(index_in_class+length_of_find_word)

            find_window = Tk()
            find_window.title("Find")
            find_window.resizable(0, 0)
            find_window.geometry("300x200")
            # Find Variable
            find_word = StringVar()
            # Find Entry Field
            find_entry = Entry(find_window, textvariable=find_word, font="Calibri 20 normal", bg="#ffffff", fg="#000000")
            find_entry.pack(anchor='center', pady=20)
            # Find Button
            find_button = Button(find_window, text="Find", font="Calibri 20 bold", bg="#ffffff", fg="#000000", width=15, command=find_in)
            find_button.pack(anchor='center', pady=5)
            find_window.mainloop()

        def find_and_replace():
            # Find & Replace Button's Function
            def find_replace():
                if (find_entry.get() == "") or (replace_entry.get() == ""):
                    messagebox.showerror("Error", "Please fill the field")
                else:
                    # checking in all fields
                    index_in_class = self.class_entry.get().find(find_entry.get())
                    index_in_first_name = self.first_entry.get().find(find_entry.get())
                    index_in_last_name = self.last_entry.get().find(find_entry.get())
                    index_in_year = self.year_entry.get().find(find_entry.get())
                    index_in_parent_name = self.parent_entry.get().find(find_entry.get())
                    index_in_address = self.address_entry.get().find(find_entry.get())
                    index_in_address_line2 = self.address_line2_entry.get().find(find_entry.get())
                    index_in_city = self.city_entry.get().find(find_entry.get())
                    index_in_pincode = self.pincode_entry.get().find(find_entry.get())
                    index_in_email = self.email_entry.get().find(find_entry.get())
                    index_in_phone = self.phone_entry.get().find(find_entry.get())
                    # Conditions
                    if (index_in_class < 0) and (index_in_first_name < 0) and (index_in_last_name < 0) and (index_in_year < 0) and (index_in_parent_name < 0) and (index_in_address < 0) and (index_in_address_line2 < 0) and (index_in_city < 0) and (index_in_pincode < 0) and (index_in_email < 0) and (index_in_phone < 0):
                        messagebox.showerror("Error", "Match Not Found!")
                    else:
                        # Class Field
                        temp_str = self.class_entry.get()
                        self.class_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.class_entry.insert(0, new)
                        # First Name Field
                        temp_str = self.first_entry.get()
                        self.first_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.first_entry.insert(0, new)
                        # Last Name Field
                        temp_str = self.last_entry.get()
                        self.last_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.last_entry.insert(0, new)
                        # Year Field
                        temp_str = self.year_entry.get()
                        self.year_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.year_entry.insert(0, new)
                        # Parent Name Field
                        temp_str = self.parent_entry.get()
                        self.parent_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.parent_entry.insert(0, new)
                        # Address Field
                        temp_str = self.address_entry.get()
                        self.address_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.address_entry.insert(0, new)
                        # Address Line 2 Field
                        temp_str = self.address_line2_entry.get()
                        self.address_line2_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.address_line2_entry.insert(0, new)
                        # City Field
                        temp_str = self.city_entry.get()
                        self.city_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.city_entry.insert(0, new)
                        # Pincode Field
                        temp_str = self.pincode_entry.get()
                        self.pincode_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.pincode_entry.insert(0, new)
                        # Email Field
                        temp_str = self.email_entry.get()
                        self.email_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.email_entry.insert(0, new)
                        # Phone Field
                        temp_str = self.phone_entry.get()
                        self.phone_entry.delete(0, 'end')
                        new = temp_str.replace(find_entry.get(), replace_entry.get())
                        self.phone_entry.insert(0, new)

            find_replace_window = Tk()
            find_replace_window.title("Find and Replace")
            find_replace_window.resizable(0, 0)
            find_replace_window.geometry("300x200")
            # Find and Replace Variables
            find_word = StringVar()
            replace_word = StringVar()
            # Find Entry Field
            find_entry = Entry(find_replace_window, textvariable=find_word, font="Calibri 20 normal", bg="#ffffff", fg="#000000")
            find_entry.pack(anchor='center', pady=20)
            # Replace Entry Field
            replace_entry = Entry(find_replace_window, textvariable=replace_word, font="Calibri 20 normal", bg="#ffffff", fg="#000000")
            replace_entry.pack(anchor='center')
            # Find Button
            find_replace_button = Button(find_replace_window, text="Find & Replace", font="Calibri 20 bold", bg="#ffffff", fg="#000000", width=15, command=find_replace)
            find_replace_button.pack(anchor='center', pady=5)
            find_replace_window.mainloop()

        def find_in_database():
            # Find Button Function
            def check_database():
                if find_entry.get() == "":
                    messagebox.showerror("Error", "Please fill the field")
                else:
                    # Class Button's Function
                    def check_in_class():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['A'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_name():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['B'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_dob():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['C'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_parent_name():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['D'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_address():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['E'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_city():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['F'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")
                        
                    def check_in_pincode():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['G'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_email():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['H'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    def check_in_phone():
                        file = load_workbook('Data.xlsx')
                        sheet = file.active
                        count = 1
                        for i in range(2, sheet.max_row+1):
                            data = sheet['I'+str(i)]
                            if find_entry.get() == data.value:
                                # print(data.value)
                                messagebox.showinfo("Information", "Match Found!")
                                break
                            else:
                                count += 1
                                continue
                        if count == sheet.max_row:
                            messagebox.showerror("Error", "Match Not Found!")

                    option_window = Tk()
                    option_window.title("Choose for Search")
                    option_window.resizable(0, 0)
                    option_window.geometry("315x165")
                    # Buttons
                        # Class Button
                    class_button = Button(option_window, text="Class", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_class)
                    class_button.place(x=10, y=10)
                        # Full Name Button
                    name_button = Button(option_window, text="Full Name", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_name)
                    name_button.place(x=70, y=10)
                        # Date of Birth Button
                    dob_button =  Button(option_window, text="Date of Birth", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_dob)
                    dob_button.place(x=173, y=10)
                        # Parent Name Button
                    parent_name_button = Button(option_window, text="Parent/Guardian Name", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_parent_name)
                    parent_name_button.place(x=10, y=60)
                        # Address Button
                    address_button = Button(option_window, text="Address", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_address)
                    address_button.place(x=225, y=60)
                        # City Button
                    city_button = Button(option_window, text="City", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_city)
                    city_button.place(x=10, y=110)
                        # Pincode Button
                    pincode_button = Button(option_window, text="Pincode", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_pincode)
                    pincode_button.place(x=60, y=110)
                        # Email Button
                    email_button = Button(option_window, text="Email", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_email)
                    email_button.place(x=145, y=110)
                        # Phone Button
                    phone_button = Button(option_window, text="Phone", font="Calibri 15 bold", bg="#ffffff", fg="#000000", command=check_in_phone)
                    phone_button.place(x=210, y=110)
                    option_window.mainloop()

            database_window = Tk()
            database_window.title("Find in Database")
            database_window.resizable(0, 0)
            database_window.geometry("300x200")
            # Find Variable
            find_word = StringVar()
            # Find Entry Field
            find_entry = Entry(database_window, textvariable=find_word, font="Calibri 20 normal", bg="#ffffff", fg="#000000")
            find_entry.pack(anchor='center', pady=20)
            # Find Button
            find_button = Button(database_window, text="Find in Database", font="Calibri 20 bold", bg="#ffffff", fg="#000000", width=15, command=check_database)
            find_button.pack(anchor='center', pady=5)
            database_window.mainloop()

            # Preference Menu's Functions
                # Theme's Functions
        def light():
            self.config(bg="#ffffff")
            self.frame_1.config(bg="#E5E5E5")
            self.head.config(bg="#E5E5E5", fg="#4BB2F9")
            self.sub_head.config(bg="#E5E5E5", fg="#4BB2F9")
            self.frame_2.config(bg="#E5E5E5")
            self.class_label.config(bg="#E5E5E5", fg="#000000")
            self.class_entry.config(bg="#ffffff")
            self.student_name_label.config(bg="#E5E5E5", fg="#000000")
            self.first_entry.config(bg="#ffffff")
            self.last_entry.config(bg="#ffffff")
            self.dob_label.config(bg="#E5E5E5", fg="#000000")
            self.year_entry.config(bg="#ffffff")
            self.parent_name_label.config(bg="#E5E5E5", fg="#000000")
            self.parent_entry.config(bg="#ffffff")
            self.address_label.config(bg="#E5E5E5", fg="#000000")
            self.address_entry.config(bg="#ffffff")
            self.address_line2_label.config(bg="#E5E5E5", fg="#000000")
            self.address_line2_entry.config(bg="#ffffff")
            self.city_label.config(bg="#E5E5E5", fg="#000000")
            self.city_entry.config(bg="#ffffff")
            self.pincode_label.config(bg="#E5E5E5", fg="#000000")
            self.pincode_entry.config(bg="#ffffff")
            self.email_label.config(bg="#E5E5E5", fg="#000000")
            self.email_entry.config(bg="#ffffff")
            self.phone_label.config(bg="#E5E5E5", fg="#000000")
            self.phone_entry.config(bg="#ffffff")
            self.clear_button.config(bg="#ffffff", fg="#4BB2F9")
            self.submit_button.config(bg="#ffffff", fg="#4BB2F9")

        def dark():
            self.config(bg="#171F24")
            self.frame_1.config(bg="#192A36")
            self.head.config(bg="#192A36", fg="#4BB2F9")
            self.sub_head.config(bg="#192A36", fg="#4BB2F9")
            self.frame_2.config(bg="#192A36")
            self.class_label.config(bg="#192A36", fg="#ffffff")
            self.class_entry.config(bg="#192A36")
            self.student_name_label.config(bg="#192A36", fg="#ffffff")
            self.first_entry.config(bg="#192A36")
            self.last_entry.config(bg="#192A36")
            self.dob_label.config(bg="#192A36", fg="#ffffff")
            self.year_entry.config(bg="#192A36")
            self.parent_name_label.config(bg="#192A36", fg="#ffffff")
            self.parent_entry.config(bg="#192A36")
            self.address_label.config(bg="#192A36", fg="#ffffff")
            self.address_entry.config(bg="#192A36")
            self.address_line2_label.config(bg="#192A36", fg="#ffffff")
            self.address_line2_entry.config(bg="#192A36")
            self.city_label.config(bg="#192A36", fg="#ffffff")
            self.city_entry.config(bg="#192A36")
            self.pincode_label.config(bg="#192A36", fg="#ffffff")
            self.pincode_entry.config(bg="#192A36")
            self.email_label.config(bg="#192A36", fg="#ffffff")
            self.email_entry.config(bg="#192A36")
            self.phone_label.config(bg="#192A36", fg="#ffffff")
            self.phone_entry.config(bg="#192A36")
            self.clear_button.config(bg="#192A36", fg="#4BB2F9")
            self.submit_button.config(bg="#192A36", fg="#4BB2F9")
            
        def black():
            self.config(bg="#121212")
            self.frame_1.config(bg="#1c1a1a")
            self.head.config(bg="#1c1a1a", fg="#ffffff")
            self.sub_head.config(bg="#1c1a1a", fg="#ffffff")
            self.frame_2.config(bg="#1c1a1a")
            self.class_label.config(bg="#1c1a1a", fg="#ffffff")
            self.class_entry.config(bg="#1c1a1a")
            self.student_name_label.config(bg="#1c1a1a", fg="#ffffff")
            self.first_entry.config(bg="#1c1a1a")
            self.last_entry.config(bg="#1c1a1a")
            self.dob_label.config(bg="#1c1a1a", fg="#ffffff")
            self.year_entry.config(bg="#1c1a1a")
            self.parent_name_label.config(bg="#1c1a1a", fg="#ffffff")
            self.parent_entry.config(bg="#1c1a1a")
            self.address_label.config(bg="#1c1a1a", fg="#ffffff")
            self.address_entry.config(bg="#1c1a1a")
            self.address_line2_label.config(bg="#1c1a1a", fg="#ffffff")
            self.address_line2_entry.config(bg="#1c1a1a")
            self.city_label.config(bg="#1c1a1a", fg="#ffffff")
            self.city_entry.config(bg="#1c1a1a")
            self.pincode_label.config(bg="#1c1a1a", fg="#ffffff")
            self.pincode_entry.config(bg="#1c1a1a")
            self.email_label.config(bg="#1c1a1a", fg="#ffffff")
            self.email_entry.config(bg="#1c1a1a")
            self.phone_label.config(bg="#1c1a1a", fg="#ffffff")
            self.phone_entry.config(bg="#1c1a1a")
            self.clear_button.config(bg="#1c1a1a", fg="#ffffff")
            self.submit_button.config(bg="#1c1a1a", fg="#ffffff")

                # Font's Functions
        def bahnschrift():
            self.head.config(font=("Bahnschrift", 30, "underline bold"))
            self.sub_head.config(font=("Bahnschrift", 15, "bold"))
            self.class_label.config(font=("Bahnschrift", 20, "bold"))
            self.student_name_label.config(font=("Bahnschrift", 20, "bold"))
            self.dob_label.config(font=("Bahnschrift", 20, "bold"))
            self.parent_name_label.config(font=("Bahnschrift", 20, "bold"))
            self.address_label.config(font=("Bahnschrift", 20, "bold"))
            self.address_line2_label.config(font=("Bahnschrift", 20, "bold"))
            self.city_label.config(font=("Bahnschrift", 20, "bold"))
            self.pincode_label.config(font=("Bahnschrift", 20, "bold"))
            self.email_label.config(font=("Bahnschrift", 20, "bold"))
            self.phone_label.config(font=("Bahnschrift", 20, "bold"))
            self.clear_button.config(font=("Bahnschrift", 20, "bold"))
            self.submit_button.config(font=("Bahnschrift", 20, "bold"))
        
        def calibri():
            self.head.config(font=("Calibri", 30, "underline bold"))
            self.sub_head.config(font=("Calibri", 15, "bold"))
            self.class_label.config(font=("Calibri", 20, "bold"))
            self.student_name_label.config(font=("Calibri", 20, "bold"))
            self.dob_label.config(font=("Calibri", 20, "bold"))
            self.parent_name_label.config(font=("Calibri", 20, "bold"))
            self.address_label.config(font=("Calibri", 20, "bold"))
            self.address_line2_label.config(font=("Calibri", 20, "bold"))
            self.city_label.config(font=("Calibri", 20, "bold"))
            self.pincode_label.config(font=("Calibri", 20, "bold"))
            self.email_label.config(font=("Calibri", 20, "bold"))
            self.phone_label.config(font=("Calibri", 20, "bold"))
            self.clear_button.config(font=("Calibri", 20, "bold"))
            self.submit_button.config(font=("Calibri", 20, "bold"))
        
        def comic_sans_ms():
            self.head.config(font=("Comic Sans MS", 30, "underline bold"))
            self.sub_head.config(font=("Comic Sans MS", 15, "bold"))
            self.class_label.config(font=("Comic Sans MS", 20, "bold"))
            self.student_name_label.config(font=("Comic Sans MS", 20, "bold"))
            self.dob_label.config(font=("Comic Sans MS", 20, "bold"))
            self.parent_name_label.config(font=("Comic Sans MS", 20, "bold"))
            self.address_label.config(font=("Comic Sans MS", 20, "bold"))
            self.address_line2_label.config(font=("Calibri", 20, "bold"))
            self.city_label.config(font=("Comic Sans MS", 20, "bold"))
            self.pincode_label.config(font=("Comic Sans MS", 20, "bold"))
            self.email_label.config(font=("Comic Sans MS", 20, "bold"))
            self.phone_label.config(font=("Comic Sans MS", 20, "bold"))
            self.clear_button.config(font=("Comic Sans MS", 20, "bold"))
            self.submit_button.config(font=("Comic Sans MS", 20, "bold"))

        def ink_free():
            self.head.config(font=("Ink Free", 30, "underline bold"))
            self.sub_head.config(font=("Ink Free", 15, "bold"))
            self.class_label.config(font=("Ink Free", 20, "bold"))
            self.student_name_label.config(font=("Ink Free", 20, "bold"))
            self.dob_label.config(font=("Ink Free", 20, "bold"))
            self.parent_name_label.config(font=("Ink Free", 20, "bold"))
            self.address_label.config(font=("Ink Free", 20, "bold"))
            self.address_line2_label.config(font=("Ink Free", 20, "bold"))
            self.city_label.config(font=("Ink Free", 20, "bold"))
            self.pincode_label.config(font=("Ink Free", 20, "bold"))
            self.email_label.config(font=("Ink Free", 20, "bold"))
            self.phone_label.config(font=("Ink Free", 20, "bold"))
            self.clear_button.config(font=("Ink Free", 20, "bold"))
            self.submit_button.config(font=("Ink Free", 20, "bold"))

        def kristen_itc():
            self.head.config(font=("Kristen ITC", 30, "underline bold"))
            self.sub_head.config(font=("Kristen ITC", 15, "bold"))
            self.class_label.config(font=("Kristen ITC", 20, "bold"))
            self.student_name_label.config(font=("Kristen ITC", 20, "bold"))
            self.dob_label.config(font=("Kristen ITC", 20, "bold"))
            self.parent_name_label.config(font=("Kristen ITC", 20, "bold"))
            self.address_label.config(font=("Kristen ITC", 20, "bold"))
            self.address_line2_label.config(font=("Kristen ITC", 20, "bold"))
            self.city_label.config(font=("Kristen ITC", 20, "bold"))
            self.pincode_label.config(font=("Kristen ITC", 20, "bold"))
            self.email_label.config(font=("Kristen ITC", 20, "bold"))
            self.phone_label.config(font=("Kristen ITC", 20, "bold"))
            self.clear_button.config(font=("Kristen ITC", 20, "bold"))
            self.submit_button.config(font=("Kristen ITC", 20, "bold"))

        def lucida_handwriting():
            self.head.config(font=("Lucida Handwriting", 30, "underline bold"))
            self.sub_head.config(font=("Lucida Handwriting", 15, "bold"))
            self.class_label.config(font=("Lucida Handwriting", 17, "bold"))
            self.student_name_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.dob_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.parent_name_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.address_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.address_line2_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.city_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.pincode_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.email_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.phone_label.config(font=("Lucida Handwriting", 20, "bold"))
            self.clear_button.config(font=("Lucida Handwriting", 20, "bold"))
            self.submit_button.config(font=("Lucida Handwriting", 20, "bold"))

        def poppins():
            self.head.config(font=("Poppins", 30, "underline bold"))
            self.sub_head.config(font=("Poppins", 15, "bold"))
            self.class_label.config(font=("Poppins", 20, "bold"))
            self.student_name_label.config(font=("Poppins", 20, "bold"))
            self.dob_label.config(font=("Poppins", 20, "bold"))
            self.parent_name_label.config(font=("Poppins", 20, "bold"))
            self.address_label.config(font=("Poppins", 20, "bold"))
            self.address_line2_label.config(font=("Poppins", 20, "bold"))
            self.city_label.config(font=("Poppins", 20, "bold"))
            self.pincode_label.config(font=("Poppins", 20, "bold"))
            self.email_label.config(font=("Poppins", 20, "bold"))
            self.phone_label.config(font=("Poppins", 20, "bold"))
            self.clear_button.config(font=("Poppins", 20, "bold"))
            self.submit_button.config(font=("Poppins", 20, "bold"))

        def randolph():
            self.head.config(font=("RANDOLPH", 30, "underline bold"))
            self.sub_head.config(font=("RANDOLPH", 15, "bold"))
            self.class_label.config(font=("RANDOLPH", 17, "bold"))
            self.student_name_label.config(font=("RANDOLPH", 20, "bold"))
            self.dob_label.config(font=("RANDOLPH", 20, "bold"))
            self.parent_name_label.config(font=("RANDOLPH", 20, "bold"))
            self.address_label.config(font=("RANDOLPH", 20, "bold"))
            self.address_line2_label.config(font=("RANDOLPH", 20, "bold"))
            self.city_label.config(font=("RANDOLPH", 20, "bold"))
            self.pincode_label.config(font=("RANDOLPH", 20, "bold"))
            self.email_label.config(font=("RANDOLPH", 20, "bold"))
            self.phone_label.config(font=("RANDOLPH", 20, "bold"))
            self.clear_button.config(font=("RANDOLPH", 20, "bold"))
            self.submit_button.config(font=("RANDOLPH", 20, "bold"))

        def times_new_roman():
            self.head.config(font=("Times New Roman", 30, "underline bold"))
            self.sub_head.config(font=("Times New Roman", 15, "bold"))
            self.class_label.config(font=("Times New Roman", 20, "bold"))
            self.student_name_label.config(font=("Times New Roman", 20, "bold"))
            self.dob_label.config(font=("Times New Roman", 20, "bold"))
            self.parent_name_label.config(font=("Times New Roman", 20, "bold"))
            self.address_label.config(font=("Times New Roman", 20, "bold"))
            self.address_line2_label.config(font=("Times New Roman", 20, "bold"))
            self.city_label.config(font=("Times New Roman", 20, "bold"))
            self.pincode_label.config(font=("Times New Roman", 20, "bold"))
            self.email_label.config(font=("Times New Roman", 20, "bold"))
            self.phone_label.config(font=("Times New Roman", 20, "bold"))
            self.clear_button.config(font=("Times New Roman", 20, "bold"))
            self.submit_button.config(font=("Times New Roman", 20, "bold"))

        def zefani_stencil():
            self.head.config(font=("Zefani Stencil", 30, "underline bold"))
            self.sub_head.config(font=("Zefani Stencil", 15, "bold"))
            self.class_label.config(font=("Zefani Stencil", 15, "bold"))
            self.student_name_label.config(font=("Zefani Stencil", 20, "bold"))
            self.dob_label.config(font=("Zefani Stencil", 20, "bold"))
            self.parent_name_label.config(font=("Zefani Stencil", 20, "bold"))
            self.address_label.config(font=("Zefani Stencil", 20, "bold"))
            self.address_line2_label.config(font=("Zefani Stencil", 20, "bold"))
            self.city_label.config(font=("Zefani Stencil", 20, "bold"))
            self.pincode_label.config(font=("Zefani Stencil", 20, "bold"))
            self.email_label.config(font=("Zefani Stencil", 20, "bold"))
            self.phone_label.config(font=("Zefani Stencil", 20, "bold"))
            self.clear_button.config(font=("Zefani Stencil", 20, "bold"))
            self.submit_button.config(font=("Zefani Stencil", 20, "bold"))

            # Help Menu's Functions
        def documentation():
            pass
        
        def social():
            social_root = Tk()
            social_root.title("Social Accounts of Developer")
            social_root.resizable(0, 0)
            social_root.geometry("420x300")
            social_root.config(bg="#000000")
            # Google Info
            google_label = Label(social_root, text="Google :- Search ", bg="#000000", fg="#ffffff", font=("Times New Roman", 20, "bold"), pady=18, padx=15)
            google_label.grid(row=0, column=3)
            ans_google = Label(social_root, text="'Ameen Sid'", bg="#000000", fg="#4BB2F9", font=("Times New Roman", 20, "bold"), pady=18)
            ans_google.grid(row=0, column=6)
            # LinkedIn Info
            linkedin_label = Label(social_root, text="LinkedIn :- ", bg="#000000", fg="#ffffff", font=("Times New Roman", 20, "bold"))
            linkedin_label.grid(row=2, column=3)
            ans_linkedin = Label(social_root, text="Ameen Sid", bg="#000000", fg="#4BB2F9", font=("Times New Roman", 20, "bold"))
            ans_linkedin.grid(row=2, column=6)
            # Twitter Info
            twitter_label = Label(social_root, text="Twitter :- ", bg="#000000", fg="#ffffff", font=("Times New Roman", 20, "bold"), pady=15)
            twitter_label.grid(row=4, column=3)
            ans_twitter = Label(social_root, text="@AmeenSid7", bg="#000000", fg="#4BB2F9", font=("Times New Roman", 20, "bold"))
            ans_twitter.grid(row=4, column=6)
            # Instagram Info
            insta_label = Label(social_root, text="Instagram :- ", bg="#000000", fg="#ffffff", font=("Times New Roman", 20, "bold"), pady=15)
            insta_label.grid(row=6, column=3)
            ans_insta = Label(social_root, text="@ameensid7", bg="#000000", fg="#4BB2F9", font=("Times New Roman", 20, "bold"))
            ans_insta.grid(row=6, column=6)
            social_root.mainloop()
        
        def about_developer():
            about_dev_root = Tk()
            about_dev_root.title("About Developer")
            about_dev_root.resizable(0, 0)
            about_dev_root.geometry("420x300")
            about_dev_root.config(bg="#ffffff")
            # Labels
            heading = Label(about_dev_root, text="Ameen Sid", bg="#ffffff", fg="#000000", font="Randolph 30 bold")
            heading.place(x=90, y=30)
            intro_1 = Label(about_dev_root, text="I am Ameen Sid, Developer of this Software.", bg="#ffffff", fg="#000000", font="Calibri 15 normal")
            intro_1.place(x=20, y=110)
            intro_2 = Label(about_dev_root, text="I am Junior Software Developer by Profession\n and Passion. I am Passionate about learning\n and evolving my self for my Technical Skills.", bg="#ffffff", fg="#000000", font="Calibri 15 normal")
            intro_2.place(x=17, y=135)
            about_dev_root.mainloop()

        def exit():
            exit()

        menuBar = Menu(self)
        self.config(menu=menuBar)
        # File Menu
        m1 = Menu(menuBar, tearoff=0)
        m1.add_command(label="New Form", command=new_form)
        m1.add_command(label="Open Record", command=open_record)
        m1.add_separator()
        m1.add_command(label="Edit Admission Details", command=edit_admission_details)
        m1.add_command(label="Cancel Admission", command=cancel_admission)
        # m1.add_command(label="Save")
        # m1.add_command(label="Save As")
        m1.add_separator()
        # m1.add_command(label="Close File")
        m1.add_command(label="Close Form", command=exit)
        m1.add_separator()
        m1.add_command(label="Exit", command=exit)
        menuBar.add_cascade(label="File", menu=m1)
        # Student Zone Menu
        std = Menu(menuBar, tearoff=0)
        std.add_command(label="Fee Status", command=fee_status)
        std.add_command(label="Fee Payment", command=fee_payment)
        std.add_separator()
        std.add_command(label="Syllabus", command=syllabus)
        menuBar.add_cascade(label="Student Zone", menu=std)
        # Edit Menu
        m2 = Menu(menuBar, tearoff=0)
        m2.add_command(label="Cut", command=lambda: self.focus_get().event_generate('<<Cut>>')) # accelerator="Ctrl+X"
        m2.add_command(label="Copy", command=lambda: self.focus_get().event_generate('<<Copy>>')) # accelerator="Ctrl+C"
        m2.add_command(label="Paste", command=lambda: self.focus_get().event_generate('<<Paste>>')) # accelerator="Ctrl+V")
        m2.add_command(label="Select All", command=lambda: self.focus_get().event_generate('<<SelectAll>>')) # accelerator="Ctrl+A")
        m2.add_separator()
            # Convert Cases - Sub Menu
        m2_1 = Menu(m2, tearoff=0)
        m2_1.add_command(label="Capitalize", command=capitalize_case)
        m2_1.add_command(label="Upper Case", command=upper_case)
        m2_1.add_command(label="Lower Case", command=lower_case)
        m2.add_cascade(label="Convert Case", menu=m2_1)
        menuBar.add_cascade(label="Edit", menu=m2)
        # Find Menu
        m3 = Menu(menuBar, tearoff=0)
        m3.add_command(label="Find", command=find_in_fields)
        m3.add_command(label="Replace", command=find_and_replace)
        m3.add_command(label="Find in Database", command=find_in_database)
        menuBar.add_cascade(label="Find", menu=m3)
        # Preference Menu
        m4 = Menu(menuBar, tearoff=0)
            # Themes - Sub Menu
        m4_1 = Menu(m4, tearoff=0)
        m4_1.add_command(label="Light", command=light)
        m4_1.add_command(label="Dark", command=dark)
        m4_1.add_command(label="Black", command=black)
        m4.add_cascade(label="Themes", menu=m4_1)
            # Fonts - Sub Menu
        m4_2 = Menu(m4, tearoff=0)
        m4_2.add_command(label="Bahnschrift", command=bahnschrift)
        m4_2.add_command(label="Calibri (Default)", command=calibri)
        m4_2.add_command(label="Comic Sans MS", command=comic_sans_ms)
        m4_2.add_command(label="Ink Free", command=ink_free)
        m4_2.add_command(label="Kristen ITC", command=kristen_itc)
        m4_2.add_command(label="Lucida Handwriting", command=lucida_handwriting)
        m4_2.add_command(label="Poppins", command=poppins)
        m4_2.add_command(label="RANDOLPH", command=randolph)
        m4_2.add_command(label="Times New Roman", command=times_new_roman)
        m4_2.add_command(label="Zefani Stencil", command=zefani_stencil)
        m4.add_cascade(label="Fonts", menu=m4_2)
        menuBar.add_cascade(label="Preference", menu=m4)
        # Help Menu
        m5 = Menu(menuBar, tearoff=0)
        m5.add_command(label="Documentation", command=documentation)
        m5.add_command(label="Social", command=social)
        m5.add_command(label="About Developer", command=about_developer)
        menuBar.add_cascade(label="Help", menu=m5)

    def form(self):
        # Functions
        def clear():
            self.class_entry.delete(0, 'end')
            self.first_entry.delete(0, 'end')
            self.last_entry.delete(0, 'end')
            self.year_entry.delete(0, 'end')
            self.parent_entry.delete(0, 'end')
            self.address_entry.delete('0', 'end')
            self.address_line2_entry.delete(0, 'end')
            self.city_entry.delete(0, 'end')
            self.pincode_entry.delete(0, 'end')
            self.email_entry.delete(0, 'end')
            self.phone_entry.delete(0, 'end')
        
        def get_details():
            # if self.class_entry.get() or self.first_entry.get() or self.year_entry.get() or self.parent_entry.get() or self.address_entry.get() or self.pincode_entry.get() or self.email_entry.get() or self.phone_entry.get() == NONE:
            if (class_var.get() == "") or (first_name_var.get() == "") or (dob_year_var.get() == "") or (parent_name_var.get() == "") or (address_var.get() == "") or (pincode_var.get() == "") or (email_var.get() == "") or (phone_var.get() == ""):
                # print("Empty")
                messagebox.showwarning("Warning", "Please fill all the fields!")
            else:
                file = load_workbook("Data.xlsx")
                sheet = file.active
                sheet.cell(column=1, row=sheet.max_row+1, value=class_var.get())
                if last_name_var.get() == "":
                    sheet.cell(column=2, row=sheet.max_row, value=first_name_var.get())
                else:
                    sheet.cell(column=2, row=sheet.max_row, value=first_name_var.get() + " " + str(last_name_var.get()))
                sheet.cell(column=3, row=sheet.max_row, value=str(days_click.get())+ " " + str(months_click.get()) + " " + str(dob_year_var.get()))
                sheet.cell(column=4, row=sheet.max_row, value=parent_name_var.get())
                if address_line2_var.get() == "":
                    sheet.cell(column=5, row=sheet.max_row, value=address_var.get())
                else:
                    sheet.cell(column=5, row=sheet.max_row, value=address_var.get() + " " + str(address_line2_var.get()))
                sheet.cell(column=6, row=sheet.max_row, value=city_var.get())
                sheet.cell(column=7, row=sheet.max_row, value=pincode_var.get())
                sheet.cell(column=8, row=sheet.max_row, value=email_var.get())
                sheet.cell(column=9, row=sheet.max_row, value=phone_var.get())
                file.save("Data.xlsx")

                # Print Data in Normal Way
                # print(class_var.get())
                # print(first_name_var.get())
                # print(last_name_var.get())
                # print(days_click.get())
                # print(months_click.get())
                # print(dob_year_var.get())
                # print(parent_name_var.get())
                # print(address_var.get())
                # print(address_line2_var.get())
                # print(city_var.get())
                # print(pincode_var.get())
                # print(email_var.get())
                # print(phone_var.get())

                # Clear Fields After Submit
                self.class_entry.delete(0, 'end')
                self.first_entry.delete(0, 'end')
                self.last_entry.delete(0, 'end')
                self.year_entry.delete(0, 'end')
                self.parent_entry.delete(0, 'end')
                self.address_entry.delete('0', 'end')
                self.address_line2_entry.delete(0, 'end')
                self.city_entry.delete(0, 'end')
                self.pincode_entry.delete(0, 'end')
                self.email_entry.delete(0, 'end')
                self.phone_entry.delete(0, 'end')

                # Showing Student Id
                id_window = Tk()
                id_window.title("Student Id")
                id_window.resizable(0, 0)
                id_window.geometry("480x200")
                id_window.config(bg="#ffffff")
                # Showing Id of Student
                Label(id_window, text="Your Admission is Conformed!", bg="#ffffff", fg="#000000", font="Calibri 25 bold").pack(anchor='center', pady=15)
                file = load_workbook("Data.xlsx")
                sheet = file.active
                student_id = "Your Student ID : 2022785" + str(sheet.max_row)
                Label(id_window, text=student_id, fg="#000000", bg="#ffffff", font="Calibri 20 bold").pack(anchor='center', pady=50)
                id_window.mainloop()

        # Define Variables
        class_var = StringVar(value="")
            # Student Name
        first_name_var = StringVar(value="")
        last_name_var = StringVar()
            # Date of Birth
        days_click = Variable()
        months_click = StringVar()
        dob_year_var = Variable(value="")
            # Parent Name
        parent_name_var = StringVar(value="")
            # Address Details
        address_var = StringVar(value="")
        address_line2_var = StringVar()
        city_var = StringVar()
        pincode_var = Variable(value="")
            # Contact Details
        email_var = StringVar(value="")
        phone_var = Variable(value="")

        # Heading Frame
        self.frame_1 = Frame(self, width=1366, height=120, bg="#E5E5E5")
        self.frame_1.pack()
        self.head = Label(self.frame_1, text="XYZ SEN. SEC. SCHOOL", fg="#4BB2F9", bg="#E5E5E5", font="Calibri 30 underline bold", pady=15)
        self.head.place(in_=self.frame_1, x=500, y=0)
        self.sub_head = Label(self.frame_1, text="School Admission Form", fg="#4BB2F9", bg="#E5E5E5", font="Calibri 15 bold", padx=0, pady=0)
        self.sub_head.place(in_=self.frame_1, x=590, y=70)

        # Form Frame
        self.frame_2 = Frame(self, width=1300, height=570, bg="#E5E5E5")
        self.frame_2.place(x=35, y=140)

        # Class
        self.class_label = Label(self.frame_2, text="Class you want to apply for*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.class_label.place(in_=self.frame_2, x=40, y=30)
        self.class_entry = Entry(self.frame_2, textvariable=class_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.class_entry.place(in_=self.frame_2, x=85, y=90)
        
        # Student Name
        self.student_name_label = Label(self.frame_2, text="Student Name*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.student_name_label.place(in_=self.frame_2, x=480, y=30)
        self.first_entry = Entry(self.frame_2, textvariable=first_name_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9", width=12)
        self.first_entry.place(in_=self.frame_2, x=525, y=90)
        self.last_entry = Entry(self.frame_2, textvariable=last_name_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9", width=12)
        self.last_entry.place(in_=self.frame_2, x=725, y=90)
        
        # Date of Birth
        self.dob_label = Label(self.frame_2, text="Date of Birth*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.dob_label.place(in_=self.frame_2, x=970, y=30)
            # Lists of Days and Months
        days_list = [i for i in range(1, 32)]
        months_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            # Set the days and Months to the Options
        days_click.set(days_list[0])
        months_click.set(months_list[0])
            # Drop Downs for Days and Months
        days_drop = OptionMenu(self.frame_2, days_click, *days_list).place(in_=self.frame_2, x=1015, y=90)
        months_drop = OptionMenu(self.frame_2, months_click, *months_list).place(in_=self.frame_2, x=1075, y=90)
        self.year_entry = Entry(self.frame_2, textvariable=dob_year_var, font="Calibri 17 normal", bg="#ffffff", fg="#4BB2F9", width=6)
        self.year_entry.place(in_=self.frame_2, x=1170, y=90)
        
        # Parent/Guardian Name
        self.parent_name_label = Label(self.frame_2, text="Parent/Guardian Name*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.parent_name_label.place(in_=self.frame_2, x=40, y=170)
        self.parent_entry = Entry(self.frame_2, textvariable=parent_name_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.parent_entry.place(in_=self.frame_2, x=85, y=230)
        
        # Address
        self.address_label = Label(self.frame_2, text="Address*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.address_label.place(in_=self.frame_2, x=480, y=170)
        self.address_entry = Entry(self.frame_2, textvariable=address_var, width=26, font="Calibri 20 normal",bg="#ffffff", fg="#4BB2F9")
        self.address_entry.place(in_=self.frame_2, x=525, y=230)
        
        # Address Line 2
        self.address_line2_label = Label(self.frame_2, text="Address Line 2", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.address_line2_label.place(in_=self.frame_2, x=970, y=170)
        self.address_line2_entry = Entry(self.frame_2, textvariable=address_line2_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9", width=17)
        self.address_line2_entry.place(in_=self.frame_2, x=1015, y=230)
        
        # City
        self.city_label = Label(self.frame_2, text="City*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.city_label.place(in_=self.frame_2, x=40, y=310)
        self.city_entry = Entry(self.frame_2, textvariable=city_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.city_entry.place(in_=self.frame_2, x=85, y=370)
        
        # Pincode
        self.pincode_label = Label(self.frame_2, text="Pincode*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.pincode_label.place(in_=self.frame_2, x=480, y=310)
        self.pincode_entry = Entry(self.frame_2, textvariable=pincode_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.pincode_entry.place(in_=self.frame_2, x=525, y=370)

        # Email
        self.email_label = Label(self.frame_2, text="Email*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.email_label.place(in_=self.frame_2, x=40, y=450)
        self.email_entry = Entry(self.frame_2, textvariable=email_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.email_entry.place(in_=self.frame_2, x=85, y=510)

        # Phone
        self.phone_label = Label(self.frame_2, text="Phone*", fg="#000000", bg="#E5E5E5", font="Calibri 20 normal")
        self.phone_label.place(in_=self.frame_2, x=480, y=450)
        self.phone_entry = Entry(self.frame_2, textvariable=phone_var, font="Calibri 20 normal", bg="#ffffff", fg="#4BB2F9")
        self.phone_entry.place(in_=self.frame_2, x=525, y=510)
        
        # Buttons
        self.clear_button = Button(self.frame_2, text="Clear", font="Calibri 20 bold", width=7, fg="#4BB2F9", bg="#ffffff", command=clear)
        self.clear_button.place(in_=self.frame_2, x=920, y=450)
        self.submit_button = Button(self.frame_2, text="Submit", font="Calibri 20 bold", width=8, fg="#4BB2F9", bg="#ffffff", command=get_details)
        self.submit_button.place(in_=self.frame_2, x=1090, y=450)

    def excel(self):
        # Excel File Create and Save
        file = pathlib.Path("Data.xlsx")
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Class"
            sheet['B1'] = "Name"
            sheet['C1'] = "Date of Birth"
            sheet['D1'] = "Parent/Guardian Name"
            sheet['E1'] = "Address"
            sheet['F1'] = "City"
            sheet['G1'] = "Pincode"
            sheet['H1'] = "Email"
            sheet['I1'] = "Phone"
            sheet['J1'] = "Fees"
            file.save("Data.xlsx")


# Main
if __name__ == '__main__':
    # Splash Screen Calls
    window = GUI()
    window.splash()
    window.frame()
    window.loading()
    window.destroy()

    # Main Screen Calls
    root = Root()
    root.menus()
    root.form()
    root.excel()
    # file = load_workbook("Data.xlsx")
    # sheet = file.active
    # print(sheet.max_row)
    root.mainloop()

    window.mainloop()
